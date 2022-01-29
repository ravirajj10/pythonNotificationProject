import openpyxl as xl
from tkinter import *
from datetime import datetime

def displaymessagefromexcel(excelstring):
    taskmsg = excelstring[0]
    win = Tk()
    now = datetime.now()
    win.geometry("600x200")
    Label(win, text= "Time-"+now.strftime('%H:%M')+"\nReminder:"+taskmsg ,font=('Helvetica bold',15)).pack(pady=50)
    win.attributes('-topmost',1)
    #win.grid_bbox(1.0,1.0,1.0,1.0)
    win.mainloop()


def process_workbook(filename):
    wb = xl.load_workbook(filename,False,False,False,False)
    sheet = wb['Sheet1']
    count = 0
    now = datetime.now()
    formatedTime = now.strftime('%H:%M')
    #print(f"currenttime {now}")
    messageDictionary = {}
    exceltimedictionary = {}
    print(f"current time {formatedTime}")
    for row in range(2, sheet.max_row+1):
        task = sheet.cell(row, 1)
        definedtime = sheet.cell(row, 2)
        remstatus = sheet.cell(row,3)
        messageDictionary[0] = task.value
        messageDictionary[1] = definedtime.value
        count = count + 1
        exceltime = str(messageDictionary[1])
        exceltimedictionary = exceltime.split(":")
        hrmin = exceltimedictionary[0]+":"+exceltimedictionary[1]
        if str(formatedTime) == hrmin:
            if remstatus.value is None:
                print("Need to remind")
                remstatus.value = "Reminded"
                displaymessagefromexcel(messageDictionary)
            else:
                print("Not Blank")
    wb.save(filename)



filename = 'reminderfortheday.xlsx'
process_workbook(filename)